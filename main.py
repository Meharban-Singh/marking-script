import openpyxl as xl
from openpyxl.styles import Font
import uuid
import sys
import re

HEADINGS_COLUMN = 'A'
FEEDBACK_COLUMN = 'D'
MAX_ROWS_TO_CHECK = 70 # efficiency
MAX_MARKS = 50
EXISTING_WORKSHEET_NAMES = ['Notes', 'Template'] # Ignores these in the Summary sheet

def main():
    
    if len(sys.argv) < 2:
        print('File name not provided')
        return 

    # Load Workbook with only data - no formulas
    wb = xl.load_workbook(sys.argv[1], data_only=True)

    output = ""
    question_comments = ""
    all_students_information = []

    # Filter out summary sheets to get accurate count for progress bar
    worksheets_to_process = [sheet for sheet in wb.worksheets if not re.match('^Summary(_.*)?$', sheet.title)]
    print(f"Processing {len(worksheets_to_process)} worksheets...")
    
    for i, sheet in enumerate(worksheets_to_process, 1):
        # Simple progress indicator
        progress_percent = (i / len(worksheets_to_process)) * 100
        sys.stdout.flush()
        sys.stdout.write(f"[{i}/{len(worksheets_to_process)}] ({progress_percent:.1f}%) Processing: {sheet.title}\n")

        # If "total" is missing in the headings column in the sheet, 
        # uncomment the line below and add the cell location where "total" should exist instead of A72
        # sheet['A72'].value = "total"
  
        # Add student name to the output
        output += "<" + sheet.title + ">\n\n"

        # Don't need to check the first heading row  
        for row in range(1, MAX_ROWS_TO_CHECK):
            # Get the heading and feedback cells in the current row
            current_heading_cell = sheet[HEADINGS_COLUMN][row]
            current_feedback_cell = sheet[FEEDBACK_COLUMN][row]

            # Check if we have a heading => the font will be bold
            if current_heading_cell.font.bold and current_heading_cell.value is not None and str(current_heading_cell.value).strip().lower() != 'total':     
                # Add question comments to the output only if it has some feedback
                if question_comments != "" and not question_comments.endswith("--\n"):
                    output += question_comments

                question_comments = "\n\n" + current_heading_cell.value + ":\n"

                # After each heading, underline with 'n' hyphens 
                for _ in range(0, len(current_heading_cell.value)):
                    question_comments += "-"

                question_comments +="\n"                


            #  If the feedback coloumn is not None, we can latee print it to file.   
            if current_feedback_cell.value is not None:
                question_comments += str(current_feedback_cell.value) + "\n"

            # Print total marks 
            if str(current_heading_cell.value).strip().lower() == 'total':
                total = sheet[chr(ord(FEEDBACK_COLUMN) - 1)][row].value
                all_students_information.append({"name": sheet.title, "total": total})

         # Add question comments to the output only if it has some feedback
        if question_comments != "" and not question_comments.endswith("--\n"):
            output += question_comments
            question_comments = ""

        # Add total to the end
        output += "\n\n\n TOTAL: " + str(total)  + "/" + str(MAX_MARKS) 
        output += "\n\n\n=====================================\n\n\n"

    create_feedback_file(output) 
    sort_worksheets(wb)
    create_summary_worksheet(wb, all_students_information)

    print("Finished")

def create_feedback_file(content : str):
    # Output the file to the same folder as the input file with file name as inputFileName_feedback.txt
    output_file_name = sys.argv[1][0: sys.argv[1].rfind(".")] + "_feedback.txt"

    try:
        # open a new file in write mode 
        with open(output_file_name, "w") as text_file:
            text_file.write(content)
        
        print(f"Feedback file '{output_file_name}' added successfully.")
    except Exception:
        print('Something went wrong while creating feedback file!')

def sort_worksheets(wb : xl.Workbook):

    print("Sorting all worksheets by name...")
    
    # Sort all worksheets by name (case-insensitive) before saving
    try:
        wb._sheets.sort(key=lambda ws: ws.title.lower())
        wb.save(sys.argv[1])
        print("Worksheets sorted successfully.")

    except Exception:
        print("Worksheets sorting failed.")

def create_summary_worksheet(wb : xl.Workbook, all_students_information : list, summary_sheet_name : str = "Summary"):

    print("Creating marks-summary worksheet...")

    # Generate a unique name if the summary sheet already exists
    if summary_sheet_name in wb.sheetnames:
        summary_sheet_name = f"Summary_{uuid.uuid4().hex[:8]}"
    
    # Create a new summary sheet
    summary_sheet = wb.create_sheet(summary_sheet_name)
    summary_sheet.append(["Name", "Total Marks"])
    
    running_total = 0
    total_students = 0

    for student_info in all_students_information:
        name = student_info.get("name", "unknown")

        if name in EXISTING_WORKSHEET_NAMES: 
            continue

        total = student_info.get("total", "unknown")
        summary_sheet.append([name, total])
        running_total += total
        total_students += 1

    summary_sheet.append([])

    avg = round(running_total / total_students, 2)
    summary_sheet.append(["Average", avg])
    
    percentage = (avg / MAX_MARKS) * 100
    summary_sheet.append(["Percentage", percentage])
    
    # Save the workbook in the same workbook
    wb.save(sys.argv[1])
    print(f"Summary sheet '{summary_sheet_name}' added successfully.")


# run main method if the current file is run
if __name__ == "__main__":
    main()