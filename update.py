# Updates all sheets in an Excel workbook by copying a specified range of cells from template sheet to all other sheets
# Useful if there was a mistake in the template that needs to be corrected across all sheets

import openpyxl as xl
from openpyxl.utils import range_boundaries
import sys 

SHEET_TO_COPY = 'Template'
SOURCE_CELL_RANGE = 'A1:B70' # for 1 cell, 'A2:A2'

def replace_sheets_with_range(source_file_path, output_file_path):
    wb_source = xl.load_workbook(source_file_path)
    wb_new = xl.load_workbook(source_file_path) # copy for output file
    
    if SHEET_TO_COPY not in wb_source.sheetnames:
        print(f"Sheet '{SHEET_TO_COPY}' does not exist.")
        return

    # Get the source sheet with the template range
    source_sheet = wb_source[SHEET_TO_COPY]
    min_col, min_row, max_col, max_row = range_boundaries(SOURCE_CELL_RANGE)
    
    for sheet_name in wb_source.sheetnames:
        if sheet_name == SHEET_TO_COPY:
            continue 
        
        current_sheet = wb_new[sheet_name]
        
        for row in source_sheet.iter_rows(min_row=min_row, 
                                          max_row=max_row, 
                                          min_col=min_col, 
                                          max_col=max_col, 
                                          values_only=False):
            for cell in row:
                new_cell = current_sheet.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                
                try:
                    if cell.font:
                        new_cell.font = xl.styles.Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color=cell.font.color,
                            scheme=cell.font.scheme
                        )
                    
                    if cell.fill:
                        new_cell.fill = xl.styles.PatternFill(
                            fill_type=cell.fill.fill_type,
                            start_color=cell.fill.start_color,
                            end_color=cell.fill.end_color
                        )
                    
                    if cell.border:
                        new_cell.border = xl.styles.Border(
                            left=xl.styles.Side(
                                border_style=cell.border.left.border_style,
                                color=cell.border.left.color
                            ),
                            right=xl.styles.Side(
                                border_style=cell.border.right.border_style,
                                color=cell.border.right.color
                            ),
                            top=xl.styles.Side(
                                border_style=cell.border.top.border_style,
                                color=cell.border.top.color
                            ),
                            bottom=xl.styles.Side(
                                border_style=cell.border.bottom.border_style,
                                color=cell.border.bottom.color
                            )
                        )
                    
                    if cell.alignment:
                        new_cell.alignment = xl.styles.Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            text_rotation=cell.alignment.text_rotation,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent
                        )
                    
                    if cell.number_format:
                        new_cell.number_format = cell.number_format
                
                except Exception as e:
                    print(f"Error copying style for cell: {e}")
    
    wb_new.save(output_file_path)
    print(f"New workbook created at {output_file_path}")

def main():
    if len(sys.argv) < 2:
        print('File name not provided') 
        return
    
    input_file = sys.argv[1]
    output_file = input_file.rsplit('.', 1)[0] + '_New.xlsx'
    
    replace_sheets_with_range(input_file, output_file)

if __name__ == "__main__":
    main()
    print('Finish')